import openpyxl as xl
from openpyxl.chart import BarChart,  PieChart,  Reference


def process_workbook(filename):
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']
    cell = sheet.cell(1, 1)
    print(cell.value)  # this will return the title of the column
    print(sheet.max_row)  # this will print the number of rows in the column
    sheet.cell(1,4).value = 'Corrected_price'
    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price

    values = Reference(sheet,
                       min_row=2,
                       max_row=sheet.max_row,
                       min_col=4,
                       max_col=4)

    chart = PieChart()
    chart.add_data(values)
    sheet.add_chart(chart, 'e2')

    wb.save(filename)


process_workbook('transaction.xlsx')